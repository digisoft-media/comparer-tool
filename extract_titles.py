import argparse
import os
import re
import sys
import time
from datetime import datetime
from typing import List, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


# ----------------------------
# Utils
# ----------------------------

def parse_number(value: str):
    """Convierte strings numéricos comunes a int/float cuando aplica."""
    if not isinstance(value, str):
        return value
    v = value.strip()
    if not v:
        return v

    if re.fullmatch(r"-?\d{1,3}(,\d{3})+(\.\d+)?", v):
        v2 = v.replace(",", "")
        return float(v2) if "." in v2 else int(v2)

    if re.fullmatch(r"-?\d+", v):
        return int(v)

    if re.fullmatch(r"-?\d+\.\d+", v):
        return float(v)

    return value


def is_title_text(txt: str) -> bool:
    return bool(txt and "User Selects" in txt)


def get_title_from_command_result(cr_element) -> Optional[str]:
    """
    Extrae el texto del título desde un WebElement (CommandResult).
    Busca:
      - div[data-testid="ansi-output"] div.ansiout
      - cualquier elemento con class que contenga 'ansiout'
      - texto que contenga "User Selects"
    """
    try:
        # 1) ansi-output -> ansiout
        ansi_outputs = cr_element.find_elements(By.CSS_SELECTOR, 'div[data-testid="ansi-output"]')
        for ao in ansi_outputs:
            ansiouts = ao.find_elements(By.CSS_SELECTOR, 'div.ansiout')
            for a in ansiouts:
                txt = (a.text or "").strip()
                if is_title_text(txt):
                    return txt
                if txt:  # a veces el título no contiene exactamente "User Selects" pero igual te sirve
                    # si quieres forzar solo User Selects, comenta esta línea:
                    if "User Selects" in txt:
                        return txt

        # 2) cualquier div.ansiout directo
        ansiouts = cr_element.find_elements(By.CSS_SELECTOR, 'div.ansiout')
        for a in ansiouts:
            txt = (a.text or "").strip()
            if is_title_text(txt):
                return txt
            if txt and "User Selects" in txt:
                return txt

        # 3) fallback: buscar cualquier nodo con texto que contenga User Selects
        # (Selenium no tiene "contains text" fácil sin XPath)
        nodes = cr_element.find_elements(By.XPATH, ".//*[contains(normalize-space(text()), 'User Selects')]")
        for n in nodes:
            txt = (n.text or "").strip()
            if is_title_text(txt):
                return txt

    except Exception:
        pass

    return None


def has_table_structure(cr_element) -> bool:
    """
    Detecta si el CommandResult parece contener una tabla.
    - class contiene 'command-result-tabs' OR
    - existe div[data-testid="datagrid.table"]
    """
    try:
        classes = (cr_element.get_attribute("class") or "")
        if "command-result-tabs" in classes:
            return True
        tables = cr_element.find_elements(By.CSS_SELECTOR, 'div[data-testid="datagrid.table"]')
        return len(tables) > 0
    except Exception:
        return False


def wait_table_rendered(wait: WebDriverWait, cr_element, timeout_sec: int = 6) -> bool:
    """
    Espera a que dentro del CommandResult exista:
      - datagrid.table
      - datagrid.grid.right
      - al menos 1 header
    """
    end = time.time() + timeout_sec
    while time.time() < end:
        try:
            datagrid = cr_element.find_elements(By.CSS_SELECTOR, 'div[data-testid="datagrid.table"]')
            if not datagrid:
                time.sleep(0.2)
                continue

            grid_right = cr_element.find_elements(By.CSS_SELECTOR, 'div[data-testid="datagrid.grid.right"]')
            if not grid_right:
                time.sleep(0.2)
                continue

            headers = cr_element.find_elements(By.CSS_SELECTOR, 'div[data-testid="datagrid.grid.right"] div[role="columnheader"]')
            if headers:
                return True
        except Exception:
            pass

        time.sleep(0.2)

    return False


def extract_table_data_from_element(cr_element) -> List[Dict[str, str]]:
    """
    Extrae tabla DIRECTAMENTE desde Selenium WebElement.
    Requisito: el bloque de tabla debe estar visible/renderizado.
    """
    table_data: List[Dict[str, str]] = []

    # datagrid + grid right
    datagrid = cr_element.find_element(By.CSS_SELECTOR, 'div[data-testid="datagrid.table"]')
    grid_right = datagrid.find_element(By.CSS_SELECTOR, 'div[data-testid="datagrid.grid.right"]')

    # headers
    headers = []
    header_elements = grid_right.find_elements(By.CSS_SELECTOR, 'div[role="columnheader"]')
    for h in header_elements:
        txt = (h.text or "").strip()
        if txt and txt != "#row_number#":
            headers.append(txt)

    if not headers:
        return table_data

    # rows (solo lo visible; en estos HTML exportados normalmente basta)
    row_elements = grid_right.find_elements(By.CSS_SELECTOR, 'div[role="row"]')
    for row in row_elements:
        # saltar header-row
        if row.find_elements(By.CSS_SELECTOR, 'div[role="columnheader"]'):
            continue

        cells = row.find_elements(By.CSS_SELECTOR, 'div[role="cell"]')
        if not cells:
            continue

        row_data = {}
        for cell in cells:
            cell_id = (cell.get_attribute("data-cell-id") or "").strip()
            value = (cell.text or "").strip()
            if not value:
                continue

            # cell_id suele ser "13_account", "13_Avg Sales", etc
            col_name = None
            if "_" in cell_id:
                _, col_name = cell_id.split("_", 1)
            else:
                # fallback por posición
                col_name = None

            if col_name and col_name in headers:
                row_data[col_name] = value
            elif col_name is None:
                # fallback por índice (último recurso)
                idx = cells.index(cell)
                if idx < len(headers):
                    row_data[headers[idx]] = value

        if row_data:
            table_data.append(row_data)

    return table_data


# ----------------------------
# Main extraction
# ----------------------------

def extract_titles_and_tables_live(driver, max_lookahead: int = 40) -> List[Dict]:
    wait = WebDriverWait(driver, 12)

    # obtener todos los CommandResult
    command_results = driver.find_elements(By.CSS_SELECTOR, 'div[data-testid="CommandResult"]')
    print(f"[INFO] CommandResult blocks: {len(command_results)}")

    results = []

    for idx, cr in enumerate(command_results, start=1):
        title = get_title_from_command_result(cr)
        if not title:
            continue

        print(f"\n[{idx}] TITLE: {title}")

        # buscar el siguiente CommandResult que tenga estructura de tabla
        table_data = []
        found_table_block = None

        for j in range(1, max_lookahead + 1):
            if idx - 1 + j >= len(command_results):
                break
            cand = command_results[idx - 1 + j]

            # si aparece otro título antes de tabla, cortamos
            cand_title = get_title_from_command_result(cand)
            if cand_title and cand_title != title:
                print(f"    [STOP] Next title reached before table (lookahead #{j}).")
                break

            if has_table_structure(cand):
                found_table_block = cand
                print(f"    [OK] Table structure found at lookahead #{j}.")
                break

        if found_table_block is not None:
            # scrollear y esperar renderizado
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", found_table_block)
            time.sleep(0.6)

            rendered = wait_table_rendered(wait, found_table_block, timeout_sec=8)
            if not rendered:
                print("    [WARN] Table did not render (virtualized / not loaded).")
            else:
                try:
                    table_data = extract_table_data_from_element(found_table_block)
                    print(f"    [OK] Rows extracted: {len(table_data)}")
                except Exception as e:
                    print(f"    [ERROR] Extract table failed: {e}")

        results.append({"title": title, "table_data": table_data})

    return results


# ----------------------------
# Excel
# ----------------------------

def save_to_excel(results: List[Dict], output_file: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    row = 1

    for item in results:
        title = item["title"]
        table = item["table_data"]

        ws.cell(row=row, column=1, value=title).font = Font(bold=True)
        row += 1

        if table:
            # columnas: unión ordenada
            columns = []
            seen = set()
            for r in table:
                for k in r.keys():
                    if k not in seen:
                        seen.add(k)
                        columns.append(k)

            # headers
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=row, column=col_idx, value=col_name)
                cell.font = Font(bold=True)
            row += 1

            # data
            for r in table:
                for col_idx, col_name in enumerate(columns, start=1):
                    v = r.get(col_name, "")
                    ws.cell(row=row, column=col_idx, value=parse_number(v))
                row += 1

        row += 1  # blank line

    # widths
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 24 if col_idx > 1 else 80

    wb.save(output_file)
    print(f"[OK] Excel saved: {output_file}")


# ----------------------------
# Selenium bootstrap
# ----------------------------

def create_driver():
    chrome_options = Options()
    # chrome_options.add_argument("--headless=new")  # si quieres headless
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--start-maximized")

    try:
        driver = webdriver.Chrome(options=chrome_options)
        return driver
    except Exception:
        # fallback webdriver-manager
        from webdriver_manager.chrome import ChromeDriverManager
        from selenium.webdriver.chrome.service import Service
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=chrome_options)
        return driver


def open_local_html(driver, html_file_path: str):
    abs_path = os.path.abspath(html_file_path)
    file_url = f"file:///{abs_path.replace(os.sep, '/')}"
    driver.get(file_url)

    wait = WebDriverWait(driver, 15)
    wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
    time.sleep(1.5)
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(0.5)


# ----------------------------
# Entrypoint
# ----------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Extrae títulos y tablas de archivos HTML de Databricks y los exporta a Excel"
    )
    parser.add_argument(
        "html_file",
        type=str,
        help="Ruta al archivo HTML a procesar"
    )
    
    args = parser.parse_args()
    html_file = args.html_file
    
    # Validar que el archivo existe
    if not os.path.exists(html_file):
        print(f"❌ Error: El archivo no existe: {html_file}")
        sys.exit(1)
    
    # Validar que es un archivo HTML
    if not html_file.lower().endswith(('.html', '.htm')):
        print(f"❌ Advertencia: El archivo no parece ser HTML: {html_file}")
    
    print(f"[INFO] Procesando archivo: {html_file}")

    driver = create_driver()
    try:
        open_local_html(driver, html_file)

        # IMPORTANTE: NO hacemos scroll global para “renderizar todo”.
        # Extraemos tabla por tabla cuando toca.
        results = extract_titles_and_tables_live(driver, max_lookahead=40)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out = f"titulos_{ts}.xlsx"
        save_to_excel(results, out)

    finally:
        try:
            driver.quit()
        except:
            pass
